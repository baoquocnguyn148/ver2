"""
=============================================================================
  DATA PIPELINE — Customer Loyalty & Revenue Dashboard
  Author : Senior DA Pipeline
  Input  : DB.xlsx  (sheet: FILE_DATA_HANDONLAB2)
  Output : data_model.xlsx  (5 sheets sẵn sàng import Power BI)
=============================================================================
  Sheets xuất ra:
    1. FACT_Sales         — bảng fact chính (mọi transaction)
    2. DIM_Customer       — thông tin khách hàng (1 row / customer)
    3. DIM_Product        — danh mục sản phẩm
    4. DIM_Geography      — địa lý (unique location)
    5. DIM_Date           — date dimension (Year × Quarter)
    6. AGG_Summary        — bảng tổng hợp KPI sẵn cho quick visual
=============================================================================
"""

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src import config
from src.utils.io_utils import DataIO

# ─────────────────────────────────────────────────────────────────────────────
# 0. CONFIG
# ─────────────────────────────────────────────────────────────────────────────
io = DataIO()
INPUT_FILE  = config.RAW_KEY if not config.LOCAL_MODE else "DB.xlsx"
OUTPUT_FILE = config.DATA_MODEL_FILE
SHEET_NAME  = "FILE_DATA_HANDONLAB2"

# Loyalty tier thứ tự từ thấp → cao (dùng để tạo rank)
LOYALTY_ORDER = {
    "Bronze": 1, "Silver": 2, "Gold": 3,
    "Elite": 4, "Platinum": 5, "VIP": 6
}

# Education thứ tự
EDU_ORDER = {
    "High School or Below": 1, "College": 2,
    "Bachelor": 3, "Master": 4, "Doctor": 5
}

# Quarter → tháng giữa (dùng để tạo sort key)
QUARTER_MONTH = {"Q1": 2, "Q2": 5, "Q3": 8, "Q4": 11}

print("=" * 65)
print("  DATA PIPELINE — Customer Loyalty & Revenue Dashboard")
print("=" * 65)


# ─────────────────────────────────────────────────────────────────────────────
# 1. LOAD RAW DATA
# ─────────────────────────────────────────────────────────────────────────────
print("\n[1/7] Loading raw data ...")
raw = io.read_raw_excel(sheet_name=SHEET_NAME)
print(f"      Loaded: {raw.shape[0]:,} rows × {raw.shape[1]} columns")


# ─────────────────────────────────────────────────────────────────────────────
# 2. CLEAN — loại bỏ / chuẩn hoá
# ─────────────────────────────────────────────────────────────────────────────
print("\n[2/7] Cleaning ...")

df = raw.copy()

# 2.1 Drop cột rác (Column1 toàn True, không có giá trị phân tích)
df.drop(columns=["Column1"], inplace=True)

# 2.2 Chuẩn hoá tên cột (bỏ khoảng trắng thừa, strip)
df.columns = df.columns.str.strip()

# 2.3 Strip whitespace ở các cột string
str_cols = df.select_dtypes(include="object").columns
for c in str_cols:
    df[c] = df[c].str.strip()

# 2.4 Income = 0 với Education = College → đây là pattern "sinh viên / không thu nhập"
#     Gắn flag thay vì xóa, để người dùng tự quyết khi filter
df["Income_Flag"] = np.where(df["Income"] == 0, "No Income Reported", "Has Income")

# 2.5 Flag năm 2020 là partial (chỉ ~2,819 records)
df["Is_Full_Year"] = np.where(df["Order Year"] == 2020, False, True)

print(f"      Dropped: Column1 (all True)")
print(f"      Flagged: {(df['Income_Flag']=='No Income Reported').sum():,} rows Income=0 → 'No Income Reported'")
print(f"      Flagged: {(~df['Is_Full_Year']).sum():,} rows Year=2020 → Is_Full_Year=False")


# ─────────────────────────────────────────────────────────────────────────────
# 3. FEATURE ENGINEERING — tạo cột mới
# ─────────────────────────────────────────────────────────────────────────────
print("\n[3/7] Feature engineering ...")

# ── 3.1 METRICS CỐT LÕI ─────────────────────────────────────────────────────

# Profit & Profit Margin
df["Profit"]         = df["Revenue"] - (df["Unit Cost"] * df["Quantity Sold"])
df["Profit_Margin_Pct"] = (df["Profit"] / df["Revenue"] * 100).round(4)

# ATV — Average Transaction Value (Revenue / Count orders)
df["ATV"] = (df["Revenue"] / df["Count"]).round(2)

# Revenue per Unit
df["Revenue_Per_Unit"] = (df["Revenue"] / df["Quantity Sold"]).round(2)

# Cost per Unit (đã có Unit Cost — giữ nguyên, thêm alias rõ ràng)
df["Total_Cost"] = df["Unit Cost"] * df["Quantity Sold"]

# Gross Contribution = Profit (alias rõ hơn cho dashboard label)
df["Gross_Contribution"] = df["Profit"]

print("      + Profit, Profit_Margin_Pct, ATV, Revenue_Per_Unit, Total_Cost")

# ── 3.2 DATE DIMENSION KEYS ──────────────────────────────────────────────────

# Year_Quarter (sort key dạng string)
df["Year_Quarter"] = df["Order Year"].astype(str) + "-" + df["Quarter"]

# Quarter số (dùng để sort)
df["Quarter_Num"] = df["Quarter"].str.extract(r"(\d)").astype(int)

# Date_Key (dùng để join DIM_Date)
df["Date_Key"] = df["Order Year"].astype(str) + df["Quarter_Num"].astype(str).str.zfill(2)

print("      + Year_Quarter, Quarter_Num, Date_Key")

# ── 3.3 CUSTOMER SEGMENTATION ────────────────────────────────────────────────

# Loyalty Tier Rank (Bronze=1 → VIP=6)
df["Loyalty_Rank"] = df["LoyaltyStatus"].map(LOYALTY_ORDER)

# Education Rank
df["Education_Rank"] = df["Education"].map(EDU_ORDER)

# Income Bucket
df["Income_Bucket"] = pd.cut(
    df["Income"],
    bins=[-1, 0, 20000, 40000, 60000, 80000, 100001],
    labels=["No Income", "<20K", "20K–40K", "40K–60K", "60K–80K", "80K+"]
)
df["Income_Bucket"] = df["Income_Bucket"].astype(str)

# Tenure Bucket (MonthsAsMember)
df["Tenure_Bucket"] = pd.cut(
    df["MonthsAsMember"],
    bins=[0, 36, 48, 60, 72],
    labels=["<3 yrs", "3–4 yrs", "4–5 yrs", "5+ yrs"]
)
df["Tenure_Bucket"] = df["Tenure_Bucket"].astype(str)

# CLV Segment (theo quantile)
df["CLV_Segment"] = pd.qcut(
    df["Customer Lifetime Value"],
    q=4,
    labels=["Low CLV", "Mid CLV", "High CLV", "Top CLV"]
)
df["CLV_Segment"] = df["CLV_Segment"].astype(str)

print("      + Loyalty_Rank, Education_Rank, Income_Bucket, Tenure_Bucket, CLV_Segment")

# ── 3.4 COUPON & MARKETING ───────────────────────────────────────────────────

# Coupon số
df["Coupon_Num"] = df["Coupon Response"].str.extract(r"(\d)").astype(int)

# Coupon Tier (1 = phổ biến nhất, 6 = hiếm)
coupon_pop = df["Coupon Response"].value_counts().rank(ascending=False).astype(int)
df["Coupon_Popularity_Rank"] = df["Coupon Response"].map(coupon_pop)

print("      + Coupon_Num, Coupon_Popularity_Rank")

# ── 3.5 GEOGRAPHY ────────────────────────────────────────────────────────────

# Region group (để map dễ hơn trong Power BI)
df["Region_Key"] = (
    df["Country"].str[:2].str.upper() + "_" +
    df["Province or State"].str[:5].str.upper().str.replace(" ", "")
)

print("      + Region_Key")


# ─────────────────────────────────────────────────────────────────────────────
# 4. BUILD STAR SCHEMA TABLES
# ─────────────────────────────────────────────────────────────────────────────
print("\n[4/7] Building star schema tables ...")

# ── DIM_Date ─────────────────────────────────────────────────────────────────
dim_date = (
    df[["Date_Key", "Order Year", "Quarter", "Quarter_Num", "Year_Quarter", "Is_Full_Year"]]
    .drop_duplicates()
    .sort_values(["Order Year", "Quarter_Num"])
    .reset_index(drop=True)
)
dim_date.rename(columns={"Order Year": "Year"}, inplace=True)

# Fix §4.2 — Thêm cột Date dạng ngày (end-of-quarter) để Power BI Time Intelligence hoạt động
Q_END_DATE = {"Q1": "-03-31", "Q2": "-06-30", "Q3": "-09-30", "Q4": "-12-31"}
dim_date["Date"] = pd.to_datetime(
    dim_date["Year"].astype(str) + dim_date["Quarter"].map(Q_END_DATE)
)
# Đặt cột Date lên đầu (sau Date_Key) để dễ nhận diện trong Power BI
cols_date = ["Date_Key", "Date", "Year", "Quarter", "Quarter_Num", "Year_Quarter", "Is_Full_Year"]
dim_date = dim_date[cols_date]

print(f"      DIM_Date       : {len(dim_date)} rows  (Date range: {dim_date['Date'].min().date()} → {dim_date['Date'].max().date()})")

# ── DIM_Product ───────────────────────────────────────────────────────────────
product_stats = (
    df.groupby("Product Line")
    .agg(
        Total_Revenue=("Revenue", "sum"),
        Total_Profit=("Profit", "sum"),
        Avg_Unit_Price=("Unit Sale Price", "mean"),
        Avg_Unit_Cost=("Unit Cost", "mean"),
        Total_Orders=("Count", "sum"),
        Total_Qty=("Quantity Sold", "sum"),
    )
    .reset_index()
)
product_stats["Avg_Profit_Margin_Pct"] = (
    product_stats["Total_Profit"] / product_stats["Total_Revenue"] * 100
).round(4)
product_stats["Product_ID"] = range(1, len(product_stats) + 1)
product_stats["Avg_Unit_Price"] = product_stats["Avg_Unit_Price"].round(2)
product_stats["Avg_Unit_Cost"]  = product_stats["Avg_Unit_Cost"].round(2)

# Fix §4.1 — Đổi tên "Product Line" → "Product_Line" cho nhất quán
product_stats.rename(columns={"Product Line": "Product_Line"}, inplace=True)

dim_product = product_stats[
    ["Product_ID", "Product_Line", "Total_Revenue", "Total_Profit",
     "Avg_Unit_Price", "Avg_Unit_Cost", "Total_Orders", "Total_Qty",
     "Avg_Profit_Margin_Pct"]
]

# Build Product_ID lookup map để dùng cho FACT_Sales
product_id_map = dim_product.set_index("Product_Line")["Product_ID"].to_dict()
print(f"      DIM_Product    : {len(dim_product)} rows  (Product_ID map built)")

# ── DIM_Geography ─────────────────────────────────────────────────────────────
# Fix Power BI relationship bug: Create a unique string key for geography
df["Geo_Key"] = df["Region_Key"] + "_" + df["City"].astype(str).str.replace(" ", "") + "_" + df["Postal code"].astype(str).str.replace(" ", "")

dim_geo = (
    df[["Geo_Key", "Region_Key", "Country", "Province or State", "City",
        "Postal code", "Latitude", "Longitude", "Location Code"]]
    .drop_duplicates(subset=["Geo_Key"])
    .reset_index(drop=True)
)
dim_geo.insert(0, "Geo_ID", range(1, len(dim_geo) + 1))
dim_geo.rename(columns={
    "Province or State": "Province_State",
    "Postal code": "Postal_Code",
    "Location Code": "Location_Type"
}, inplace=True)
print(f"      DIM_Geography  : {len(dim_geo)} rows")

# ── DIM_Customer ──────────────────────────────────────────────────────────────
# Lấy 1 row / customer — dùng max CLV và max Revenue làm representative
dim_customer = (
    df.sort_values("Customer Lifetime Value", ascending=False)
    .drop_duplicates(subset=["Loyalty#"])
    .reset_index(drop=True)
)
dim_customer = dim_customer[[
    "Loyalty#", "First Name", "Last Name", "Customer Name",
    "Gender", "Education", "Education_Rank",
    "Marital Status", "Income", "Income_Flag", "Income_Bucket",
    "MonthsAsMember", "Tenure_Bucket",
    "LoyaltyStatus", "Loyalty_Rank",
    "Customer Lifetime Value", "CLV_Segment"
]].copy()
dim_customer.rename(columns={
    "Loyalty#": "Customer_ID",
    "First Name": "First_Name",
    "Last Name": "Last_Name",
    "Customer Name": "Full_Name",
    "Marital Status": "Marital_Status",
    "MonthsAsMember": "Months_As_Member",
    "Customer Lifetime Value": "CLV"
}, inplace=True)
print(f"      DIM_Customer   : {len(dim_customer):,} rows  ({dim_customer['Customer_ID'].nunique():,} unique customers)")

# ── FACT_Sales ─────────────────────────────────────────────────────────────────
fact_sales = df[[
    # Keys
    "Loyalty#", "Date_Key", "Product Line", "Region_Key", "Geo_Key",
    # Transaction
    "Coupon Response", "Coupon_Num", "Coupon_Popularity_Rank",
    "Count", "Quantity Sold", "Unit Sale Price", "Unit Cost",
    # Metrics
    "Revenue", "Total_Cost", "Profit", "Profit_Margin_Pct",
    "ATV", "Revenue_Per_Unit", "Gross_Contribution",
    # Customer features at transaction time
    "LoyaltyStatus", "Loyalty_Rank",
    "Income_Bucket", "Tenure_Bucket", "CLV_Segment",
    # Time features
    "Order Year", "Quarter", "Quarter_Num", "Year_Quarter", "Is_Full_Year"
]].copy()

fact_sales.rename(columns={
    "Loyalty#": "Customer_ID",
    "Product Line": "Product_Line",
    "Coupon Response": "Coupon_Response",
    "Quantity Sold": "Quantity_Sold",
    "Unit Sale Price": "Unit_Sale_Price",
    "Unit Cost": "Unit_Cost",
    "Order Year": "Year",
    "Quarter": "Quarter",
    "Count": "Order_Count"
}, inplace=True)

# Fix §4.1 — Thêm Product_ID (integer FK) vào FACT_Sales để join với DIM_Product bằng số
fact_sales["Product_ID"] = fact_sales["Product_Line"].map(product_id_map)

# Thêm Surrogate Key cho fact
fact_sales.insert(0, "Fact_ID", range(1, len(fact_sales) + 1))
print(f"      FACT_Sales     : {len(fact_sales):,} rows  (Product_ID FK added)")


# ─────────────────────────────────────────────────────────────────────────────
# 5. AGG SUMMARY — bảng tổng hợp sẵn cho Quick Visual / Power BI Import
# ─────────────────────────────────────────────────────────────────────────────
print("\n[5/7] Building aggregation summaries ...")

def agg_base(group_cols):
    return df.groupby(group_cols, observed=True).agg(
        Total_Revenue=("Revenue", "sum"),
        Total_Profit=("Profit", "sum"),
        Total_Orders=("Count", "sum"),
        Total_Qty=("Quantity Sold", "sum"),
        Avg_CLV=("Customer Lifetime Value", "mean"),
        Unique_Customers=("Loyalty#", "nunique"),
    ).reset_index()

# 5.1 Revenue by Year × Product Line
agg_year_product = agg_base(["Order Year", "Product Line"])
agg_year_product["Profit_Margin_Pct"] = (
    agg_year_product["Total_Profit"] / agg_year_product["Total_Revenue"] * 100
).round(2)

# 5.2 Revenue by Country × Year
agg_country_year = agg_base(["Country", "Order Year"])
agg_country_year["Profit_Margin_Pct"] = (
    agg_country_year["Total_Profit"] / agg_country_year["Total_Revenue"] * 100
).round(2)

# 5.3 Customer profile by LoyaltyStatus
agg_loyalty = df.groupby("LoyaltyStatus", observed=True).agg(
    Customer_Count=("Loyalty#", "nunique"),
    Total_Revenue=("Revenue", "sum"),
    Total_Profit=("Profit", "sum"),
    Avg_Revenue=("Revenue", "mean"),
    Avg_CLV=("Customer Lifetime Value", "mean"),
    Avg_Income=("Income", lambda x: x[x>0].mean()),  # exclude 0-income
    Avg_Months=("MonthsAsMember", "mean"),
).reset_index()
agg_loyalty["Loyalty_Rank"] = agg_loyalty["LoyaltyStatus"].map(LOYALTY_ORDER)
agg_loyalty["Pct_Revenue"] = (
    agg_loyalty["Total_Revenue"] / agg_loyalty["Total_Revenue"].sum() * 100
).round(2)
agg_loyalty = agg_loyalty.sort_values("Loyalty_Rank")
agg_loyalty["Avg_Revenue"]  = agg_loyalty["Avg_Revenue"].round(2)
agg_loyalty["Avg_CLV"]      = agg_loyalty["Avg_CLV"].round(2)
agg_loyalty["Avg_Income"]   = agg_loyalty["Avg_Income"].round(2)
agg_loyalty["Avg_Months"]   = agg_loyalty["Avg_Months"].round(1)

# 5.4 Education × Product Line
agg_edu_product = agg_base(["Education", "Product Line"])
agg_edu_product["Education_Rank"] = agg_edu_product["Education"].map(EDU_ORDER)
agg_edu_product = agg_edu_product.sort_values("Education_Rank")

# 5.5 Gender × Product Line
agg_gender_product = agg_base(["Gender", "Product Line"])

# 5.6 Coupon effectiveness
agg_coupon = df.groupby("Coupon Response", observed=True).agg(
    Usage_Count=("Count", "sum"),
    Total_Revenue=("Revenue", "sum"),
    Avg_Revenue=("Revenue", "mean"),
    Total_Profit=("Profit", "sum"),
).reset_index()
agg_coupon["Profit_Margin_Pct"] = (
    agg_coupon["Total_Profit"] / agg_coupon["Total_Revenue"] * 100
).round(2)
agg_coupon["Revenue_Per_Use"] = (
    agg_coupon["Total_Revenue"] / agg_coupon["Usage_Count"]
).round(2)
agg_coupon["Pct_Usage"] = (
    agg_coupon["Usage_Count"] / agg_coupon["Usage_Count"].sum() * 100
).round(2)
agg_coupon = agg_coupon.sort_values("Usage_Count", ascending=False)

# 5.7 Overall KPI summary
kpi = {
    "Metric": [
        "Total Revenue ($)",
        "Total Profit ($)",
        "Overall Profit Margin (%)",
        "Total Transactions",
        "Unique Customers",
        "Avg CLV ($)",
        "Avg Revenue per Transaction ($)",
        "Avg Revenue per Customer ($)",
        "Top Product Line (Revenue)",
        "Top Country (Revenue)",
        "Top LoyaltyStatus (Revenue)",
    ],
    "Value": [
        f"{df['Revenue'].sum():,.0f}",
        f"{df['Profit'].sum():,.0f}",
        f"{df['Profit'].sum() / df['Revenue'].sum() * 100:.2f}",
        f"{df['Count'].sum():,.0f}",
        f"{df['Loyalty#'].nunique():,}",
        f"{df['Customer Lifetime Value'].mean():,.2f}",
        f"{df['ATV'].mean():,.2f}",
        f"{df.groupby('Loyalty#')['Revenue'].sum().mean():,.2f}",
        df.groupby('Product Line')['Revenue'].sum().idxmax(),
        df.groupby('Country')['Revenue'].sum().idxmax(),
        df.groupby('LoyaltyStatus')['Revenue'].sum().idxmax(),
    ]
}
agg_kpi = pd.DataFrame(kpi)

print("      Built 7 aggregation tables")


# ─────────────────────────────────────────────────────────────────────────────
# 6. EXPORT TO EXCEL (multi-sheet)
# ─────────────────────────────────────────────────────────────────────────────
print("\n[6/7] Exporting to Excel ...")

sheet_map = {
    "FACT_Sales":        fact_sales,
    "DIM_Customer":      dim_customer,
    "DIM_Product":       dim_product,
    "DIM_Geography":     dim_geo,
    "DIM_Date":          dim_date,
    "AGG_KPI":           agg_kpi,
    "AGG_Year_Product":  agg_year_product,
    "AGG_Country_Year":  agg_country_year,
    "AGG_Loyalty":       agg_loyalty,
    "AGG_Edu_Product":   agg_edu_product,
    "AGG_Gender_Product":agg_gender_product,
    "AGG_Coupon":        agg_coupon,
}

print("      Exporting curated Parquet tables ...")
fact_sales_curated = fact_sales.copy()
fact_sales_curated["partition_year"] = fact_sales_curated["Year"].astype(str)

dim_customer_curated = dim_customer.copy()
dim_customer_curated["partition_loyalty_status"] = dim_customer_curated["LoyaltyStatus"].astype(str)

dim_date_curated = dim_date.copy()
dim_date_curated["partition_year"] = dim_date_curated["Year"].astype(str)

dim_geo_curated = dim_geo.copy()
dim_geo_curated["partition_country"] = dim_geo_curated["Country"].astype(str)

curated_tables = {
    "fact_sales":     (fact_sales_curated,    ["partition_year"]),
    "dim_customer":   (dim_customer_curated,  ["partition_loyalty_status"]),
    "dim_product":    (dim_product,           []),
    "dim_date":       (dim_date_curated,      ["partition_year"]),
    "dim_geography":  (dim_geo_curated,       ["partition_country"]),
}
for table_name, (table_df, partition_cols) in curated_tables.items():
    location = io.save_parquet(table_name, table_df, partition_cols=partition_cols)
    print(f"      Curated: {table_name:<14} -> {location}")

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    for sheet_name, data in sheet_map.items():
        data.to_excel(writer, sheet_name=sheet_name, index=False)

# ─────────────────────────────────────────────────────────────────────────────
# 7. FORMATTING — header styling
# ─────────────────────────────────────────────────────────────────────────────
print("\n[7/7] Applying formatting ...")

HEADER_COLORS = {
    "FACT_Sales":         "0C447C",
    "DIM_Customer":       "0F6E56",
    "DIM_Product":        "854F0B",
    "DIM_Geography":      "3C3489",
    "DIM_Date":           "3B6D11",
    "AGG_KPI":            "444441",
    "AGG_Year_Product":   "185FA5",
    "AGG_Country_Year":   "185FA5",
    "AGG_Loyalty":        "185FA5",
    "AGG_Edu_Product":    "185FA5",
    "AGG_Gender_Product": "185FA5",
    "AGG_Coupon":         "185FA5",
}

wb = load_workbook(OUTPUT_FILE)

thin = Side(style="thin", color="D0D0D0")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

for sheet_name, data in sheet_map.items():
    ws = wb[sheet_name]
    hdr_color = HEADER_COLORS.get(sheet_name, "185FA5")

    # Header row
    for cell in ws[1]:
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", fgColor=hdr_color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border

    # Auto-fit column width
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1), start=1):
        max_len = 0
        for cell in col_cells:
            try:
                max_len = max(max_len, len(str(cell.value or "")))
            except Exception:
                pass
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze top row
    ws.freeze_panes = "A2"

    # Alternating row fill
    light_fill = PatternFill("solid", fgColor="F5F7FA")
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for cell in row:
            if row_idx % 2 == 0:
                cell.fill = light_fill
            cell.font   = Font(name="Arial", size=9)
            cell.border = border

# Tab colors
TAB_COLORS = {
    "FACT_Sales": "0C447C",
    "DIM_Customer": "0F6E56", "DIM_Product": "854F0B",
    "DIM_Geography": "3C3489", "DIM_Date": "3B6D11",
}
for sn, color in TAB_COLORS.items():
    wb[sn].sheet_properties.tabColor = color

wb.save(OUTPUT_FILE)
excel_output_location = str(OUTPUT_FILE)
if not config.LOCAL_MODE:
    output_key = f"{config.OUTPUT_PREFIX.strip('/')}/data_model.xlsx"
    excel_output_location = io.upload_file(OUTPUT_FILE, output_key)

# ─────────────────────────────────────────────────────────────────────────────
# DONE — Print summary
# ─────────────────────────────────────────────────────────────────────────────
print("\n" + "=" * 65)
print("  ✅  PIPELINE COMPLETE")
print("=" * 65)
print(f"  Output : {excel_output_location}")
print(f"\n  Sheets exported:")
for name, data in sheet_map.items():
    print(f"    {'●':1} {name:<25} {len(data):>7,} rows  ×  {data.shape[1]:>2} cols")

print(f"""
  New columns added to FACT_Sales:
    Profit, Profit_Margin_Pct, ATV, Revenue_Per_Unit,
    Total_Cost, Gross_Contribution, Year_Quarter,
    Quarter_Num, Date_Key, Loyalty_Rank, Education_Rank,
    Income_Bucket, Tenure_Bucket, CLV_Segment,
    Coupon_Num, Coupon_Popularity_Rank, Region_Key,
    Income_Flag, Is_Full_Year

  Power BI import order (Star Schema):
    1. DIM_Date      → join FACT_Sales on Date_Key        (1:N)
    2. DIM_Customer  → join FACT_Sales on Customer_ID     (1:N)
    3. DIM_Product   → join FACT_Sales on Product_ID      (1:N) ← integer FK
    4. DIM_Geography → join FACT_Sales on Region_Key      (1:N)
    5. FACT_Sales (center table)
    6. AGG_* tables  → nguồn quick charts, KHÔNG relationship

  Fixes applied:
    [§4.1] DIM_Product[Product_Line] renamed, Product_ID added to FACT_Sales
    [§4.2] DIM_Date[Date] added (end-of-quarter dates for Time Intelligence)
    [§0]   INPUT/OUTPUT paths changed to relative paths
""")
